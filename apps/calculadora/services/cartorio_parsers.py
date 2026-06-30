"""Extração e aplicação automática de tabelas cartorárias a partir das fontes.

Fluxo "totalmente automático" (escolha do usuário):
- O monitoramento (cartorio_fontes.verificar_fonte_cartorio) detecta mudança na
  fonte oficial e chama `aplicar_da_fonte`.
- Se houver um parser registrado para a UF, extraímos as faixas de escritura e
  registro, elas passam por GUARDAS DE SANIDADE e, se válidas, gravamos novas
  CartorioTabela vigentes (status `validada`), aposentando as anteriores.
- Se NÃO houver parser, ou o parse falhar nas guardas, devolvemos
  ``aplicado=False`` e o monitoramento cai no comportamento antigo: apenas
  sinaliza as tabelas para revisão manual. Isso evita publicar um valor
  "plausível-porém-errado" quando o PDF do TJ vier quebrado.

Cada parser recebe ``(payload, fetch)``:
- ``payload``: conteúdo já baixado da URL monitorada (dict com ``bytes`` e
  ``content_type``). Para a maioria das UFs a URL monitorada já é o PDF da
  tabela, então o parser usa ``payload["bytes"]`` direto.
- ``fetch``: função ``fetch(url) -> bytes | None`` para baixar fontes extras
  (ex.: SP, cuja escritura vem de uma planilha separada).
E devolve ``{"escritura": [(limite, valor), ...], "registro": [...]}`` ou None.

ATENÇÃO: as URLs em ``FONTE_TABELA_URL`` e as registradas no seed
(``FONTES_URL``) são específicas do exercício (têm o ano no caminho). Quando o
TJ publicar a tabela do ano seguinte numa nova URL, a fonte antiga passa a
retornar 404 (vira evento de erro) — basta atualizar a URL.
"""
import logging
import re
from datetime import date
from decimal import Decimal

import requests
from django.conf import settings
from django.db import transaction
from django.utils import timezone

logger = logging.getLogger(__name__)

# Guardas de sanidade -------------------------------------------------------
MIN_FAIXAS = 5            # tabela real tem várias faixas; menos que isso é suspeito
VALOR_MIN = Decimal("10")        # emolumento plausível mínimo
VALOR_MAX = Decimal("500000")    # teto plausível (acima disso é erro de parse)
LIMITE_MAX = Decimal("100000000")  # R$ 100 mi como limite superior plausível

_NUM = r"[\d.]+,\d{2}"


class FaixasInvalidasError(ValueError):
    """Faixas extraídas não passaram nas guardas de sanidade."""


def _num(s):
    return Decimal(s.replace(".", "").replace(",", "."))


def validar_faixas(faixas):
    """Valida uma lista de faixas (limite_superior|None, valor).

    Regras: quantidade mínima, valores estritamente crescentes e dentro de uma
    faixa plausível, limites crescentes, e última faixa "sem limite" (None).
    Devolve (ok: bool, motivo: str).
    """
    if not faixas or len(faixas) < MIN_FAIXAS:
        return False, f"poucas faixas ({len(faixas) if faixas else 0} < {MIN_FAIXAS})"

    limites = [f[0] for f in faixas]
    valores = [Decimal(str(f[1])) for f in faixas]

    if limites[-1] is not None:
        return False, "última faixa não é 'sem limite' (None)"
    if any(l is None for l in limites[:-1]):
        return False, "faixa intermediária sem limite superior"

    limites_num = [Decimal(str(l)) for l in limites[:-1]]
    for i in range(1, len(limites_num)):
        if limites_num[i] <= limites_num[i - 1]:
            return False, f"limites não crescentes em #{i}"
    if limites_num and limites_num[-1] > LIMITE_MAX:
        return False, "limite superior implausível"

    for i, v in enumerate(valores):
        if v < VALOR_MIN or v > VALOR_MAX:
            return False, f"valor fora do plausível em #{i}: {v}"
        if i > 0 and v < valores[i - 1]:
            return False, f"valores não crescentes em #{i}"

    return True, "ok"


# Helpers de extração -------------------------------------------------------
def _abrir_pdf(conteudo_bytes):
    try:
        import fitz  # PyMuPDF
    except ImportError:  # pragma: no cover
        logger.warning("PyMuPDF indisponível; extração automática desabilitada")
        return None
    try:
        return fitz.open(stream=conteudo_bytes, filetype="pdf")
    except Exception as exc:  # pragma: no cover - PDF inválido
        logger.warning("Falha ao abrir PDF para extração: %s", exc)
        return None


def _texto_pdf(conteudo_bytes):
    doc = _abrir_pdf(conteudo_bytes)
    if doc is None:
        return ""
    return "\n".join(page.get_text() for page in doc)


def _fechar_inf(faixas):
    """Garante última faixa sem limite (None), repetindo o último valor."""
    if not faixas:
        return faixas
    if faixas[-1][0] is None:
        return faixas
    return faixas + [(None, faixas[-1][1])]


def extrair_faixas_pdf(conteudo_bytes):
    """Extrai faixas de um PDF com estrutura simples 'faixa -> valor único'.

    Cobre linhas do tipo ``... até R$ X ... R$ Y`` / ``de R$ A a R$ B ... R$ Y``
    e ``a partir de R$ Z ... R$ Y`` (última). Devolve a maior sequência
    monotônica encontrada, ou None.
    """
    texto = _texto_pdf(conteudo_bytes)
    if not texto:
        return None
    faixas = []
    prev = Decimal("-1")
    padrao = re.compile(
        r"(?:a partir de\s*R?\$?\s*(?P<ap>%s))"
        r"|(?:(?:até|de [\d.]+,\d{2}\s*a)\s*R?\$?\s*(?P<up>%s))" % (_NUM, _NUM),
        re.IGNORECASE,
    )
    valor_re = re.compile(_NUM)
    for linha in texto.split("\n"):
        m = padrao.search(linha)
        if not m:
            continue
        is_inf = bool(m.group("ap"))
        limite = None if is_inf else _num(m.group("up"))
        nums = valor_re.findall(linha)
        if len(nums) < 2 and not is_inf:
            continue
        valor = _num(nums[-1])
        if valor >= prev:
            faixas.append((limite, valor))
            prev = valor
        if is_inf:
            break
    return faixas or None


# Parsers por UF ------------------------------------------------------------
# URLs oficiais das tabelas (ano-específicas — ver nota no topo do módulo).
FONTE_TABELA_URL = {
    "BA": "https://www.tjba.jus.br/extrajudicial/wp-content/uploads/2025/12/TABELA-DE-CUSTAS-TJBA-2026-C-DECRETO.pdf",
    "MG": "https://www8.tjmg.jus.br/institucional/at/pdf/cpo86642025.pdf",
    "GO": "https://www.tjgo.jus.br/files/2026/01%20-%20Janeiro/Provimento%20Conjunto%20179-2026.pdf",
    "PE": "https://portal.tjpe.jus.br/documents/d/portal/ato-1556-tab-emolumentos-dj390-2025-pdf",
    "RS": "https://colegioregistralrs.org.br/img/emolumentos/39_emolumentos.pdf",
    "SC": "https://risantoamaro.com.br/wp-content/uploads/2026/02/Tabela-de-Emolumentos-2026-1.pdf",
    "SP_REGISTRO": "https://www.registrodeimoveis.org.br/intranet/arquivos/upload/geral/888826_upload_de_arquivos_geral_20251226_174008_f6561d12-df91-4fd0-ad28-a8204b8b9d03.pdf",
    "SP_NOTAS": "https://cnbsp.org.br/wp-content/uploads/2025/12/TABELA-DE-EMOLUMENTOS-BASE-PARA-2026-Capital.xlsx",
}


def _bytes_pdf(payload):
    """Devolve os bytes do PDF já baixado pela fonte monitorada (ou None).

    A fonte monitorada deve apontar para o PDF canônico da tabela (ver
    FONTE_TABELA_URL / seed). Não fazemos download da fonte primária aqui para
    manter o parse determinístico e sem rede oculta; fontes secundárias (ex.: a
    planilha de Notas do SP) são buscadas explicitamente via ``fetch``.
    """
    if not payload:
        return None
    ct = (payload.get("content_type") or "").lower()
    if "pdf" in ct and payload.get("bytes"):
        return payload["bytes"]
    return None


def _ba_tabela(texto, prefixo):
    """Extrai a tabela do TJ-BA ancorando no código do ato (Item I = 01xxx
    Notas; Item VII = 07xxx Registro). Cada faixa é um bloco de linhas terminado
    pelo código; o valor é o número imediatamente antes do código."""
    out = []
    pat = r"(?:Até|\na)\s*\n(%s)\s*\n(%s)\s*\n(%s\d{3})" % (_NUM, _NUM, prefixo)
    for m in re.finditer(pat, texto):
        up, v = _num(m.group(1)), _num(m.group(2))
        if not out or v >= out[-1][1]:
            out.append((up, v))
    mi = re.search(r"A partir de\s*\n%s\s*\n(%s)\s*\n%s\d{3}" % (_NUM, _NUM, prefixo), texto)
    if mi:
        out.append((None, _num(mi.group(1))))
    return _fechar_inf(out)


def parse_ba(payload, fetch):
    conteudo = _bytes_pdf(payload)
    if not conteudo:
        return None
    texto = _texto_pdf(conteudo)
    esc = _ba_tabela(texto, "01")  # Item I - Notas (escritura)
    reg = _ba_tabela(texto, "07")  # Item VII - Registro de Imóveis
    if len(esc) < MIN_FAIXAS or len(reg) < MIN_FAIXAS:
        return None
    return {"escritura": esc, "registro": reg}


def parse_mg(payload, fetch):
    conteudo = _bytes_pdf(payload)
    if not conteudo:
        return None
    texto = _texto_pdf(conteudo)
    i = texto.find("e) Escritura pública, instrumento particular e título judicial, com conteúdo financeiro")
    if i < 0:
        return None
    linhas = [l.strip() for l in texto[i:i + 2600].split("\n")]
    faixas = []
    j = 0
    while j < len(linhas):
        m = re.match(r"(?:de [\d.]+,\d{2} )?at[ée] (%s)" % _NUM, linhas[j])
        if m:
            nums, k = [], j + 1
            while k < len(linhas) and len(nums) < 3:
                if re.fullmatch(_NUM, linhas[k]):
                    nums.append(_num(linhas[k])); k += 1
                elif linhas[k] == "":
                    k += 1
                else:
                    break
            if len(nums) == 3 and (not faixas or nums[2] >= faixas[-1][1]):
                faixas.append((_num(m.group(1)), nums[2]))
            elif len(nums) == 3:
                break
            j = k
        else:
            j += 1
    faixas = _fechar_inf(faixas)
    return {"escritura": faixas, "registro": faixas} if len(faixas) >= MIN_FAIXAS else None


def parse_sc(payload, fetch):
    conteudo = _bytes_pdf(payload)
    if not conteudo:
        return None
    linhas = [l.strip() for l in _texto_pdf(conteudo).split("\n") if l.strip()]
    faixas = []
    for j, l in enumerate(linhas):
        m = re.match(r"2\.2\.\d+\.\s*(?:até|de)\s*(%s)(?:\s*a\s*(%s))?" % (_NUM, _NUM), l)
        if not m:
            continue
        up = _num(m.group(2) or m.group(1))
        nums, k = [], j + 1
        while k < len(linhas) and len(nums) < 4:
            if re.fullmatch(_NUM, linhas[k]):
                nums.append(_num(linhas[k])); k += 1
            else:
                break
        if len(nums) >= 4 and (not faixas or nums[3] >= faixas[-1][1]):
            faixas.append((up, nums[3]))
    faixas = _fechar_inf(sorted(set(faixas)))
    # SC: esta fonte só traz Registro; usamos como aproximação p/ escritura.
    return {"escritura": faixas, "registro": faixas} if len(faixas) >= MIN_FAIXAS else None


def _go_tabela(seg):
    out = []
    for m in re.finditer(r"até\s*R\$\s*R?\$?\s*(%s)[^\n]*\n\s*R\$\s*(%s)" % (_NUM, _NUM), seg):
        up, v = _num(m.group(1)), _num(m.group(2))
        if not out or v >= out[-1][1]:
            out.append((up, v))
    fim = re.search(r"acima de[^\n]*?(%s)[^\n]*\n\s*R\$\s*(%s)" % (_NUM, _NUM), seg, re.I)
    if fim:
        out.append((None, _num(fim.group(2))))
    return _fechar_inf(out)


def parse_go(payload, fetch):
    conteudo = _bytes_pdf(payload)
    if not conteudo:
        return None
    texto = _texto_pdf(conteudo)
    ie, ir = texto.find("63 –  Escritura"), texto.find("76 - Registro")
    if ie < 0 or ir < 0:
        return None
    je, jr = texto.find("\n64", ie + 10), texto.find("\n77", ir + 10)
    esc = _go_tabela(texto[ie:je if je > 0 else ie + 5000])
    reg = _go_tabela(texto[ir:jr if jr > 0 else ir + 5000])
    if len(esc) < MIN_FAIXAS or len(reg) < MIN_FAIXAS:
        return None
    return {"escritura": esc, "registro": reg}


def _pe_seq(seg, cap=None):
    out = []
    for m in re.finditer(r"(?:A|Até|até) R\$ ?(%s)\s*\n\s*R\$ ?(%s)" % (_NUM, _NUM), seg):
        up, v = _num(m.group(1)), _num(m.group(2))
        if not out or v >= out[-1][1] - Decimal("0.001"):
            out.append((up, v))
        else:
            break
    if cap is not None:
        out.append((None, cap))
    return _fechar_inf(out)


def parse_pe(payload, fetch):
    conteudo = _bytes_pdf(payload)
    if not conteudo:
        return None
    texto = _texto_pdf(conteudo)
    d, e = texto.find("TABELA 'D'"), texto.find("TABELA 'E'")
    ivp = texto.find("IV -  Registro de quaisquer atos", e)
    if d < 0 or e < 0 or ivp < 0:
        return None
    mcap = re.search(r"Emolumentos máximos:\s*\n?\s*R\$ ?(%s)" % _NUM, texto[d:e])
    esc = _pe_seq(texto[d:e], cap=_num(mcap.group(1)) if mcap else None)
    mv = re.search(r"\n(V|VI|VII)\s*[-–]", texto[ivp + 10:])
    endr = (mv.start() + ivp + 10) if mv else len(texto)
    mfim = re.search(r"A partir de R\$ ?%s\s*\n\s*R\$ ?(%s)" % (_NUM, _NUM), texto[ivp:endr])
    reg = _pe_seq(texto[ivp:endr], cap=_num(mfim.group(1)) if mfim else None)
    if len(esc) < MIN_FAIXAS or len(reg) < MIN_FAIXAS:
        return None
    return {"escritura": esc, "registro": reg}


def _rs_seq(seg, prefixos):
    out = []
    pat = r"(?:%s) R\$ ?(%s)[.\s]*R\$ ?(%s)" % ("|".join(prefixos), _NUM, _NUM)
    for m in re.finditer(pat, seg):
        up, v = _num(m.group(1)), _num(m.group(2))
        if not out or v >= out[-1][1]:
            out.append((up, v))
        else:
            break
    return out


def parse_rs(payload, fetch):
    conteudo = _bytes_pdf(payload)
    if not conteudo:
        return None
    texto = _texto_pdf(conteudo)
    ie = texto.find("i) outras escrituras com conteúdo financeiro")
    esc = _rs_seq(texto[ie:ie + 4000] if ie >= 0 else "", ["até"])
    mfe = re.search(r"acima de R\$ ?%s[.\s]*R\$ ?(%s)" % (_NUM, _NUM), texto[ie:ie + 4000]) if ie >= 0 else None
    if mfe:
        esc.append((None, _num(mfe.group(1))))
    ir = texto.lower().find("registro de imóveis")
    segr = texto[ir:ir + 4000] if ir >= 0 else ""
    reg = _rs_seq(segr, ["de valor até", "acima até", "De valor até", "Acima até"])
    mfr = re.search(r"[Aa]cima de R\$ ?%s[.\s]*R\$ ?(%s)" % (_NUM, _NUM), segr)
    if mfr:
        reg.append((None, _num(mfr.group(1))))
    esc, reg = _fechar_inf(esc), _fechar_inf(reg)
    if len(esc) < MIN_FAIXAS or len(reg) < MIN_FAIXAS:
        return None
    return {"escritura": esc, "registro": reg}


def _sp_registro_pdf(conteudo, total_xmin=465):
    """SP Registro = Tabela III.B (pág. 1), coluna TOTAL, por coordenadas."""
    from collections import defaultdict
    doc = _abrir_pdf(conteudo)
    if doc is None or doc.page_count < 2:
        return []
    words = doc[1].get_text("words")
    lines = defaultdict(list)
    for w in words:
        lines[round(w[1])].append(w)
    brackets, totals = [], []
    for y in sorted(lines):
        toks = sorted(lines[y], key=lambda w: w[0])
        m = re.search(r"até R\$ (%s)" % _NUM, " ".join(t[4] for t in toks))
        if m:
            brackets.append((toks[0][1], _num(m.group(1))))
    for w in words:
        if w[0] >= total_xmin and re.fullmatch(_NUM, w[4]):
            totals.append((w[1], _num(w[4])))
    tem_inf = "superior a" in doc[1].get_text().lower()
    out = []
    for by, up in brackets:
        cand = sorted((abs(ty - by), v) for ty, v in totals)
        if cand and cand[0][0] <= 8:
            out.append((up, cand[0][1]))
    if tem_inf and totals:
        out.append((None, max(v for _, v in totals)))
    return _fechar_inf(out)


def _sp_notas_xlsx(conteudo):
    """SP Escritura = planilha CNB/SP 'Capital', col. E (faixa) e P (TOTAL)."""
    try:
        import io
        import openpyxl
    except ImportError:  # pragma: no cover
        return []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception:  # pragma: no cover
        return []
    nome = next((s for s in wb.sheetnames if "CAPITAL" in s.upper()), wb.sheetnames[0])
    ws = wb[nome]
    out = []
    for r in range(6, 60):
        up, tot = ws.cell(r, 5).value, ws.cell(r, 16).value
        if not isinstance(tot, (int, float)):
            continue
        lim = None if up in (0, None) else Decimal(str(up))
        v = Decimal(str(round(tot, 2)))
        if not out or v >= out[-1][1]:
            out.append((lim, v))
        if lim is None:
            break
    return _fechar_inf(out)


def parse_sp(payload, fetch):
    pdf = _bytes_pdf(payload)
    reg = _sp_registro_pdf(pdf) if pdf else []
    xlsx = fetch(FONTE_TABELA_URL["SP_NOTAS"])
    esc = _sp_notas_xlsx(xlsx) if xlsx else []
    if len(esc) < MIN_FAIXAS or len(reg) < MIN_FAIXAS:
        return None
    return {"escritura": esc, "registro": reg}


# UF -> parser(payload, fetch). UFs sem parser caem na revisão manual.
PARSERS = {
    "BA": parse_ba,
    "MG": parse_mg,
    "SC": parse_sc,
    "GO": parse_go,
    "PE": parse_pe,
    "RS": parse_rs,
    "SP": parse_sp,
}


# Aplicação no banco --------------------------------------------------------
def _faixas_para_objs(tabela, faixas, modelo_faixa):
    objs = []
    for ordem, (limite, valor) in enumerate(faixas, start=1):
        objs.append(
            modelo_faixa(
                tabela=tabela,
                ordem=ordem,
                limite_superior=None if limite is None else Decimal(str(limite)),
                valor=Decimal(str(valor)),
            )
        )
    return objs


@transaction.atomic
def aplicar_tabela(uf, tipo, faixas, *, fonte_nome, fonte_url, fundamento="",
                   ano=None, observacao="", validar=True):
    """Grava uma nova CartorioTabela vigente para (uf, tipo) com as faixas dadas.

    Aposenta (substituida + vigente_fim) as tabelas vigentes anteriores do mesmo
    par (uf, tipo) e cria a nova como vigente a partir de hoje, status `validada`
    (modo totalmente automático). Levanta FaixasInvalidasError se as guardas
    reprovarem.
    """
    from apps.calculadora.models import CartorioFaixa, CartorioTabela

    if validar:
        ok, motivo = validar_faixas(faixas)
        if not ok:
            raise FaixasInvalidasError(motivo)

    hoje = date.today()
    ano = ano or hoje.year
    ontem = hoje.fromordinal(hoje.toordinal() - 1)

    anteriores = CartorioTabela.objects.filter(uf=uf, tipo=tipo, ativo=True).exclude(
        vigente_inicio=hoje
    )
    for ant in anteriores:
        ant.status = CartorioTabela.STATUS_SUBSTITUIDA
        if ant.vigente_fim is None or ant.vigente_fim >= hoje:
            ant.vigente_fim = ontem
        ant.save(update_fields=["status", "vigente_fim", "atualizado_em"])

    tabela, _ = CartorioTabela.objects.update_or_create(
        uf=uf,
        ano=ano,
        tipo=tipo,
        vigente_inicio=hoje,
        defaults={
            "fonte_nome": fonte_nome,
            "fonte_url": fonte_url or "",
            "fundamento": fundamento,
            "observacoes": observacao,
            "status": CartorioTabela.STATUS_VALIDADA,
            "ativo": True,
            "vigente_fim": None,
            "conferido_em": timezone.now(),
        },
    )
    tabela.faixas.all().delete()
    CartorioFaixa.objects.bulk_create(_faixas_para_objs(tabela, faixas, CartorioFaixa))
    return tabela


# Orquestração --------------------------------------------------------------
_USER_AGENT = "LoteAlvo/1.0 (+https://lotealvo.up.railway.app; parser cartorario)"


def _baixar(url, session=None):
    session = session or requests.Session()
    try:
        resp = session.get(url, headers={"User-Agent": _USER_AGENT}, timeout=40)
        if resp.status_code >= 400:
            return None
        return resp.content
    except requests.RequestException as exc:  # pragma: no cover - rede
        logger.warning("Falha ao baixar fonte de tabela %s: %s", url, exc)
        return None


def aplicar_da_fonte(fonte, payload, *, session=None, evento=None):
    """Tenta extrair e aplicar automaticamente as tabelas de uma fonte alterada.

    Devolve dict com ``aplicado`` (bool), ``tabelas`` (tipos aplicados) e
    ``motivo`` (quando não aplicado). Nunca levanta — falha vira aplicado=False
    para o monitoramento cair no fluxo de revisão manual.
    """
    resultado = {"aplicado": False, "tabelas": [], "motivo": ""}

    if not getattr(settings, "CARTORIO_AUTO_APLICAR_TABELAS", True):
        resultado["motivo"] = "aplicação automática desabilitada por configuração"
        return resultado

    parser = PARSERS.get((fonte.uf or "").upper())
    if not parser:
        resultado["motivo"] = f"sem parser registrado para UF {fonte.uf}"
        return resultado

    def fetch(url):
        return _baixar(url, session=session)

    try:
        tabelas = parser(payload, fetch)
    except Exception as exc:  # pragma: no cover - parser defensivo
        logger.warning("Parser cartório falhou para %s: %s", fonte, exc)
        resultado["motivo"] = f"erro no parser: {exc}"
        return resultado

    if not tabelas:
        resultado["motivo"] = "parser não retornou faixas"
        return resultado

    aplicaveis = []
    for tipo in ("escritura", "registro"):
        faixas = tabelas.get(tipo)
        if not faixas:
            continue
        ok, motivo = validar_faixas(faixas)
        if not ok:
            logger.info("Faixas reprovadas (%s/%s): %s", fonte.uf, tipo, motivo)
            resultado["motivo"] = f"{tipo}: {motivo}"
            return resultado  # tudo-ou-nada: não aplica parcial
        aplicaveis.append((tipo, faixas))

    if not aplicaveis:
        resultado["motivo"] = "nenhuma faixa de escritura/registro encontrada"
        return resultado

    obs = "Tabela extraída e aplicada automaticamente pelo monitoramento de fontes."
    if evento is not None:
        obs += f" Evento de fonte #{evento.pk}."
    for tipo, faixas in aplicaveis:
        aplicar_tabela(
            fonte.uf,
            tipo,
            faixas,
            fonte_nome=f"Fonte oficial TJ-{fonte.uf} (automático)",
            fonte_url=fonte.url,
            observacao=obs,
            validar=False,  # já validamos acima
        )
        resultado["tabelas"].append(tipo)

    resultado["aplicado"] = True
    return resultado
